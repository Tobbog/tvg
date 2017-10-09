import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "tvg.settings")

import django
django.setup()

import openpyxl
from agcs.models import Companies, Years

#indicate number of rows
num_rows = 6343
#open excel file
wb = openpyxl.load_workbook("agcs/test.xlsx")
#indicate sheet name
Daten = wb.get_sheet_by_name("Daten")

#Create array with available years
years = ["2010", "2011", "2012", "2013", "2014", "2015", "2016"] #7 years!
numYears = len(years)
#for the range of available companies:
for i in range(2, num_rows):
    #if there is an ESG:
    if isinstance(Daten.cell(row = i, column = 95).value, (int, float, complex)):
        #add the base values to the company
        c = Companies()
        c.name = Daten.cell(row=i, column=2).value
        c.nation = Daten.cell(row=i, column=3).value
        c.DSCD = Daten.cell(row=i, column=4).value
        c.ESG = Daten.cell(row=i, column=95).value
        #if there is the ENV Value:
        if isinstance(Daten.cell(row = i, column = 108).value, (int, float, complex)):
            #add the Env value
            c.ENV = Daten.cell(row=i, column=108).value
        #if SOC:
        if isinstance(Daten.cell(row = i, column = 121).value, (int, float, complex)):
            #add SOC
            c.SOC = Daten.cell(row=i, column=121).value
        #if CGV:
        if isinstance(Daten.cell(row = i, column = 134).value, (int, float, complex)):
            #add CGV
            c.CGV = Daten.cell(row=i, column=134).value
        #if QUAL:
        if isinstance(Daten.cell(row = i, column = 147).value, (int, float, complex)):
            #add QUAL
            c.QUAL = Daten.cell(row=i, column=147).value
        #save instance
        c.save()
        #for the range of available years:
        for j in range(numYears):
            #if the available year has an ESG:
            if isinstance(Daten.cell(row = i, column = 89+j).value, (int, float, complex)):
                #add the base and ESG values to the year and save it
                yr = Years()
                yr.company = c
                yr.year = years[j]
                yr.ESG = Daten.cell(row = i, column = 89+j).value
                #Repeat for the four ESG Factors:
                if isinstance(Daten.cell(row = i, column = 102+j).value, (int, float, complex)):
                    yr.ENV = Daten.cell(row=i, column=102+j).value
                if isinstance(Daten.cell(row = i, column = 115+j).value, (int, float, complex)):
                    yr.SOC = Daten.cell(row=i, column=115+j).value
                if isinstance(Daten.cell(row = i, column = 128+j).value, (int, float, complex)):
                    yr.CGV = Daten.cell(row=i, column=128+j).value
                if isinstance(Daten.cell(row = i, column = 141+j).value, (int, float, complex)):
                    yr.QUAL = Daten.cell(row=i, column=141+j).value
                yr.save()
