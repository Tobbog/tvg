import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "tvg.settings")

import django
django.setup()

import openpyxl
from agcs.models import Companies

#indicate number of rows
num_rows = 6343
#open excel file
wb = openpyxl.load_workbook("agcs/test.xlsx")
#indicate sheet name
Daten = wb.get_sheet_by_name("Daten")

for i in range(2, num_rows):
    c = Companies()
    #check if the ESG data is an empty cell
    if isinstance(Daten.cell(row = i, column = 95).value, (int, float, complex)):
        c.name = Daten.cell(row=i, column=2).value
        c.nation = Daten.cell(row=i, column=3).value
        c.DSCD = Daten.cell(row=i, column=4).value
        c.ESG = Daten.cell(row=i, column=95).value
    if isinstance(Daten.cell(row = i, column = 108).value, (int, float, complex)):
        c.ENV = Daten.cell(row=i, column=108).value
    if isinstance(Daten.cell(row = i, column = 121).value, (int, float, complex)):
        c.SOC = Daten.cell(row=i, column=121).value
    if isinstance(Daten.cell(row = i, column = 134).value, (int, float, complex)):
        c.CGV = Daten.cell(row=i, column=134).value
    if isinstance(Daten.cell(row = i, column = 147).value, (int, float, complex)):
        c.QUAL = Daten.cell(row=i, column=147).value
        c.save()
